#!/usr/bin/env python3
"""
Test de validación con el Excel real — Mide tiempos por fase.

Iteración 1: Cache vacía → todo va a Nominatim
Iteración 2: Cache poblada → todo se resuelve desde memoria

Uso:
  python test_validation_flow.py
"""

import asyncio
import time
import sys
import os

# Asegurar que el import de app funcione
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))


async def main():
    # ── Parsear el Excel real (replica la lógica de Flutter/ExcelService) ──
    import openpyxl, re
    wb = openpyxl.load_workbook("real_excel.xlsx", data_only=True)

    all_addresses = []
    all_names = []

    # Patrones heurísticos (iguales a los de Flutter _detectFromData)
    street_re = re.compile(
        r'(calle|c/|c\.|avda|avenida|plaza|pza|camino|ctra|carretera|paseo|ronda|travesía|trav)',
        re.IGNORECASE,
    )
    locality_re = re.compile(
        r'(posadas|córdoba|cordoba|sevilla|jaén|granada|málaga|almería|huelva|cádiz)',
        re.IGNORECASE,
    )
    header_kw = [
        'nombre', 'dirección', 'direccion', 'localidad', 'kgs',
        'bultos', 'bult', 'albaran', 'expedicion', 'c.p.',
        'srv', 'hr', 'reem', 'deb', 'adu', 'rcs',
    ]
    # Regex para la hoja embebida (Table 4)
    cp_pattern = re.compile(
        r'14\d{3}\s{2,}(.+?)\s{3,}(.+?)\s{3,}(Posadas(?:[,\s]*(?:CO|Córdoba|Cordoba)?)?)',
        re.IGNORECASE,
    )

    def _cell_str(val):
        if val is None:
            return ""
        return str(val).strip()

    def _is_header_row(cells):
        matches = sum(1 for c in cells if any(k in _cell_str(c).lower() for k in header_kw))
        return matches >= 2

    def _is_last_sheet(sheet):
        for row in sheet.iter_rows(values_only=True):
            for c in (row or []):
                if c and 'total env' in str(c).lower():
                    return True
        return False

    def _detect_from_header(cells):
        """Detect columns from header row (same as Flutter _detectFromHeader)."""
        nombre_idx = dir_idx = loc_idx = bultos_idx = -1
        for i, c in enumerate(cells):
            h = _cell_str(c).lower()
            if nombre_idx < 0 and ('nombre' in h or h == 'name' or 'cliente' in h):
                nombre_idx = i
            if dir_idx < 0 and ('direc' in h or 'address' in h or 'calle' in h):
                dir_idx = i
            if loc_idx < 0 and ('localidad' in h or 'ciudad' in h or 'poblacion' in h or 'población' in h):
                loc_idx = i
            if bultos_idx < 0 and ('bult' in h or 'paquetes' in h or 'packages' in h):
                bultos_idx = i
        if dir_idx < 0:
            return None
        if nombre_idx < 0:
            nombre_idx = dir_idx - 1 if dir_idx > 0 else 0
        return (nombre_idx, dir_idx, loc_idx, bultos_idx)

    def _detect_from_data(rows):
        """Detect columns heuristically from data (same as Flutter _detectFromData)."""
        if not rows:
            return None
        num_cols = max(len(r) for r in rows)
        if num_cols < 5:
            return None
        sample = rows[:10]
        dir_idx = loc_idx = -1
        for col in range(num_cols):
            street_matches = sum(1 for r in sample if col < len(r) and street_re.search(_cell_str(r[col])))
            loc_matches = sum(1 for r in sample if col < len(r) and locality_re.search(_cell_str(r[col])))
            if street_matches > len(sample) * 0.3 and dir_idx < 0:
                dir_idx = col
            if loc_matches > len(sample) * 0.3 and loc_idx < 0:
                loc_idx = col
        # Fallback: if we found locality but no address, address = locality - 1
        if loc_idx >= 0 and dir_idx < 0 and loc_idx > 0:
            dir_idx = loc_idx - 1
        if dir_idx < 0:
            return None
        nombre_idx = dir_idx - 1 if dir_idx > 0 else 0
        return (nombre_idx, dir_idx, loc_idx, -1)

    def _parse_embedded(sheet):
        """Parse last sheet with embedded text (Table 4)."""
        entries = []
        for row in sheet.iter_rows(values_only=True):
            for c in (row or []):
                text = _cell_str(c)
                if not text:
                    continue
                for line in text.split('\n'):
                    line = line.strip()
                    if not line or 'total env' in line.lower():
                        continue
                    m = cp_pattern.search(line)
                    if m:
                        nombre = m.group(1).strip()
                        direccion = m.group(2).strip()
                        if direccion:
                            full = direccion if 'posadas' in direccion.lower() else f"{direccion}, Posadas"
                            entries.append((nombre, full))
        return entries

    # ── Procesar cada hoja ──
    sheets = wb.worksheets
    for si, sheet in enumerate(sheets):
        is_first = (si == 0)

        # ¿Es la hoja de totales/embebida (última)?
        if _is_last_sheet(sheet):
            for nombre, addr in _parse_embedded(sheet):
                all_addresses.append(addr)
                all_names.append(nombre)
            print(f"   Hoja {si} ({sheet.title}): embebida → {len(_parse_embedded(sheet))} entradas")
            continue

        # Leer todas las filas
        all_rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        if not all_rows:
            continue

        # Detectar cabecera y columnas
        has_header = is_first and _is_header_row(all_rows[0])
        if has_header:
            col_map = _detect_from_header(all_rows[0])
        else:
            col_map = _detect_from_data(all_rows)

        if col_map is None:
            print(f"   ⚠️  Hoja {si} ({sheet.title}): no se detectaron columnas, saltando")
            continue

        nombre_idx, dir_idx, loc_idx, bultos_idx = col_map
        print(f"   Hoja {si} ({sheet.title}): nombre=col{nombre_idx}, dir=col{dir_idx}, loc=col{loc_idx}")

        data_start = 1 if has_header else 0
        count = 0
        for row in all_rows[data_start:]:
            nombre = _cell_str(row[nombre_idx]) if nombre_idx >= 0 and nombre_idx < len(row) else ""
            direccion = _cell_str(row[dir_idx]) if dir_idx >= 0 and dir_idx < len(row) else ""
            localidad = _cell_str(row[loc_idx]) if loc_idx >= 0 and loc_idx < len(row) else ""

            if not nombre and not direccion:
                continue
            first_cell = _cell_str(row[0]).lower() if row else ""
            if 'total' in first_cell or 'firma' in first_cell:
                continue
            if 'total env' in nombre.lower():
                continue

            full_address = direccion
            if localidad and localidad.lower() not in direccion.lower():
                full_address = f"{direccion}, {localidad}"

            if full_address.strip():
                all_addresses.append(full_address.strip())
                all_names.append(nombre)
                count += 1

        print(f"             → {count} direcciones extraídas")

    print(f"📂 Excel parseado: {len(all_addresses)} direcciones totales")
    unique_addrs = list(set(all_addresses))
    print(f"   {len(unique_addrs)} direcciones únicas")

    # ── Inicializar DB + cache en memoria ──
    from app.services.street_db import (
        get_db, load_memory_cache,
        _mem_aliases, _mem_virtuals, _mem_cache,
    )
    await get_db()

    # ══════════════════════════════════════════════
    #  BORRAR cache para simular primera vez
    # ══════════════════════════════════════════════
    db = await get_db()
    await db.execute("DELETE FROM geocode_cache")
    await db.execute("DELETE FROM alias")
    await db.execute("DELETE FROM street_virtual")
    await db.commit()
    await load_memory_cache()
    print(f"\n🗑️  Cache limpiada: {len(_mem_aliases)} alias, {len(_mem_virtuals)} virtual, {len(_mem_cache)} cache")

    # ── Importar el endpoint directamente ──
    from app.routers.validation import validation_start, StartRequest

    # ══════════════════════════════════════════════
    #  ITERACIÓN 1: Sin cache (todo a Nominatim)
    # ══════════════════════════════════════════════
    print("\n" + "="*60)
    print("  ITERACIÓN 1: SIN CACHE (todo va a Nominatim)")
    print("="*60)

    req1 = StartRequest(addresses=all_addresses, client_names=all_names)
    t1_start = time.time()
    resp1 = await validation_start(req1)
    t1_total = (time.time() - t1_start) * 1000

    print(f"\n📊 Resultado Iteración 1:")
    print(f"   Total paradas: {resp1.total_stops}")
    print(f"   Total calles:  {resp1.total_streets}")
    print(f"   ✅ OK:          {resp1.ok_streets}")
    print(f"   ❌ Problema:    {resp1.problem_streets}")
    print(f"   ⏱️  Tiempo total: {t1_total:.0f}ms")

    print(f"\n   Fases:")
    for p in resp1.phases:
        print(f"   Fase {p.phase} ({p.name}): {p.elapsed_ms:.1f}ms — {p.detail}")

    # Mostrar calles OK con su fuente
    ok_streets = [s for s in resp1.streets if s.status == "ok"]
    problem_streets = [s for s in resp1.streets if s.status == "problem"]

    print(f"\n   ✅ Calles resueltas ({len(ok_streets)}):")
    for s in ok_streets[:10]:
        print(f"      {s.street_display:40s} → {s.canonical_name:30s} [{s.source}] conf={s.confidence}")
    if len(ok_streets) > 10:
        print(f"      ... y {len(ok_streets) - 10} más")

    print(f"\n   ❌ Calles con problema ({len(problem_streets)}):")
    for s in problem_streets:
        print(f"      {s.street_display:40s} — {s.reason}")

    # Ver estado de la cache tras iteración 1
    await load_memory_cache()
    print(f"\n💾 Cache tras iteración 1: {len(_mem_aliases)} alias, {len(_mem_virtuals)} virtual, {len(_mem_cache)} cache")

    # ══════════════════════════════════════════════
    #  ITERACIÓN 2: Con cache poblada
    # ══════════════════════════════════════════════
    print("\n" + "="*60)
    print("  ITERACIÓN 2: CON CACHE (debería ser instantánea)")
    print("="*60)

    req2 = StartRequest(addresses=all_addresses, client_names=all_names)
    t2_start = time.time()
    resp2 = await validation_start(req2)
    t2_total = (time.time() - t2_start) * 1000

    print(f"\n📊 Resultado Iteración 2:")
    print(f"   Total paradas: {resp2.total_stops}")
    print(f"   Total calles:  {resp2.total_streets}")
    print(f"   ✅ OK:          {resp2.ok_streets}")
    print(f"   ❌ Problema:    {resp2.problem_streets}")
    print(f"   ⏱️  Tiempo total: {t2_total:.0f}ms")

    print(f"\n   Fases:")
    for p in resp2.phases:
        print(f"   Fase {p.phase} ({p.name}): {p.elapsed_ms:.1f}ms — {p.detail}")

    # ══════════════════════════════════════════════
    #  COMPARATIVA
    # ══════════════════════════════════════════════
    print("\n" + "="*60)
    print("  COMPARATIVA")
    print("="*60)
    speedup = t1_total / t2_total if t2_total > 0 else float('inf')
    print(f"   Iteración 1 (sin cache): {t1_total:.0f}ms")
    print(f"   Iteración 2 (con cache): {t2_total:.0f}ms")
    print(f"   Speedup: {speedup:.0f}x más rápido")
    print(f"   Ahorro: {t1_total - t2_total:.0f}ms")


if __name__ == "__main__":
    asyncio.run(main())
