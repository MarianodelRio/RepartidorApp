#!/usr/bin/env python3
"""
Convierte excel_final.xlsx → paradas.csv

Formato de salida (CSV limpio):
    cliente,direccion,ciudad

- Extrae Nombre, Dirección y Localidad de todas las hojas del Excel
- La última hoja (con "Total envíos") se parsea como texto embebido
- No agrupa duplicados: cada fila = 1 paquete
- El CSV resultante es el que importa la app Flutter directamente
"""

import csv
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl


def cell_str(cell) -> str:
    """Devuelve el valor de una celda como string limpio."""
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def is_header_row(row_values: list[str]) -> bool:
    """Detecta si una fila es cabecera mirando palabras clave."""
    keywords = {"nombre", "dirección", "direccion", "localidad", "bultos",
                "bult", "kgs", "expedicion", "albaran", "c.p.", "srv"}
    matches = sum(
        1 for v in row_values
        if any(k in v.lower() for k in keywords)
    )
    return matches >= 2


def detect_columns(header: list[str]) -> dict[str, int]:
    """Detecta índices de Nombre, Dirección, Localidad desde la cabecera."""
    cols: dict[str, int] = {"nombre": -1, "direccion": -1, "localidad": -1}
    for i, h in enumerate(header):
        hl = h.lower().strip()
        if cols["nombre"] < 0 and ("nombre" in hl or "cliente" in hl):
            cols["nombre"] = i
        if cols["direccion"] < 0 and ("direcc" in hl or "calle" in hl):
            cols["direccion"] = i
        if cols["localidad"] < 0 and ("localidad" in hl or "ciudad" in hl or "poblac" in hl):
            cols["localidad"] = i
    return cols


def is_total_sheet(ws) -> bool:
    """La última hoja contiene 'Total envíos'."""
    for row in ws.iter_rows(max_row=min(ws.max_row, 5)):
        for cell in row:
            if cell.value and "total env" in str(cell.value).lower():
                return True
    return False


def parse_embedded_text(ws) -> list[dict]:
    """Parsea la hoja de texto embebido (última hoja)."""
    entries = []
    cp_pattern = re.compile(
        r"14\d{3}\s{2,}(.+?)\s{3,}(.+?)\s{3,}"
        r"(Posadas(?:[,\s]*(?:CO|Córdoba|Cordoba)?)?)",
        re.IGNORECASE,
    )
    for row in ws.iter_rows():
        for cell in row:
            text = str(cell.value or "")
            if not text.strip():
                continue
            for line in text.split("\n"):
                line = line.strip()
                if not line or "total env" in line.lower():
                    continue
                m = cp_pattern.search(line)
                if m:
                    nombre = m.group(1).strip()
                    direccion = m.group(2).strip()
                    ciudad = m.group(3).strip()
                    # Limpiar ciudad
                    ciudad = re.sub(r",?\s*(CO|Córdoba|Cordoba)\s*$", "", ciudad, flags=re.I).strip()
                    if not ciudad:
                        ciudad = "Posadas"
                    entries.append({
                        "cliente": nombre,
                        "direccion": direccion,
                        "ciudad": ciudad,
                    })
    return entries


def read_row(ws, row_idx: int) -> list[str]:
    """Lee una fila como lista de strings."""
    return [cell_str(ws.cell(row=row_idx, column=c))
            for c in range(1, ws.max_column + 1)]


def parse_data_sheet(ws, cols: dict[str, int] | None, has_header: bool) -> list[dict]:
    """Parsea una hoja de datos estructurada."""
    entries = []
    
    if cols is None and has_header:
        header = read_row(ws, 1)
        cols = detect_columns(header)
    
    if cols is None:
        # Intentar detectar de los datos con heurística
        # En este Excel, las hojas sin cabecera pueden tener columnas desplazadas
        # Buscamos columnas que contengan calles y localidades
        return _parse_sheet_heuristic(ws)
    
    start_row = 2 if has_header else 1
    
    for r in range(start_row, ws.max_row + 1):
        row = read_row(ws, r)
        
        nombre = row[cols["nombre"]] if 0 <= cols["nombre"] < len(row) else ""
        direccion = row[cols["direccion"]] if 0 <= cols["direccion"] < len(row) else ""
        localidad = row[cols["localidad"]] if 0 <= cols["localidad"] < len(row) else ""
        
        if not nombre and not direccion:
            continue
        # Saltar filas de totales
        first = row[0].lower() if row else ""
        if "total" in first or "firma" in first:
            continue
        if "total env" in nombre.lower():
            continue
        
        if not direccion.strip():
            continue
        
        # Limpiar localidad
        localidad = localidad.strip()
        if not localidad:
            localidad = "Posadas"
        # Normalizar variaciones
        localidad = re.sub(r",?\s*(Andalucia|CO|Córdoba|Cordoba)\s*$", "", localidad, flags=re.I).strip()
        if not localidad:
            localidad = "Posadas"
        
        entries.append({
            "cliente": nombre.strip(),
            "direccion": direccion.strip(),
            "ciudad": localidad,
        })
    
    return entries


def _parse_sheet_heuristic(ws) -> list[dict]:
    """Detecta columnas heurísticamente para hojas sin cabecera."""
    street_re = re.compile(
        r"(calle|c/|avda|avenida|plaza|camino|ctra|paseo|ronda|prol)",
        re.IGNORECASE,
    )
    locality_re = re.compile(
        r"(posadas|córdoba|cordoba|sevilla)",
        re.IGNORECASE,
    )
    
    # Muestrear las primeras filas para detectar columnas
    samples = []
    for r in range(1, min(ws.max_row + 1, 11)):
        samples.append(read_row(ws, r))
    
    if not samples:
        return []
    
    num_cols = max(len(row) for row in samples)
    dir_col = -1
    loc_col = -1
    
    for c in range(num_cols):
        street_hits = sum(1 for row in samples if c < len(row) and street_re.search(row[c]))
        loc_hits = sum(1 for row in samples if c < len(row) and locality_re.search(row[c]))
        if street_hits >= len(samples) * 0.2 and dir_col < 0:
            dir_col = c
        if loc_hits >= len(samples) * 0.3 and loc_col < 0:
            loc_col = c
    
    if dir_col < 0 and loc_col >= 0:
        dir_col = loc_col - 1
    if dir_col < 0:
        return []
    
    name_col = dir_col - 1 if dir_col > 0 else 0
    
    cols = {"nombre": name_col, "direccion": dir_col, "localidad": loc_col}
    return parse_data_sheet(ws, cols, has_header=False)


def main():
    excel_path = Path(__file__).parent / "excel_final.xlsx"
    csv_path = Path(__file__).parent / "paradas.csv"
    
    if not excel_path.exists():
        print(f"❌ No se encontró {excel_path}")
        sys.exit(1)
    
    wb = openpyxl.load_workbook(str(excel_path))
    sheet_names = wb.sheetnames
    all_entries: list[dict] = []
    
    # Detectar columnas desde la primera hoja (tiene cabecera)
    first_ws = wb[sheet_names[0]]
    header = read_row(first_ws, 1)
    cols = detect_columns(header) if is_header_row(header) else None
    
    for idx, name in enumerate(sheet_names):
        ws = wb[name]
        is_first = idx == 0
        
        if is_total_sheet(ws):
            entries = parse_embedded_text(ws)
            print(f"  📋 {name}: {len(entries)} entregas (texto embebido)")
        else:
            entries = parse_data_sheet(
                ws,
                cols=cols if is_first else None,
                has_header=is_first,
            )
            print(f"  📋 {name}: {len(entries)} entregas")
        
        all_entries.extend(entries)
    
    # Escribir CSV
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["cliente", "direccion", "ciudad"])
        writer.writeheader()
        writer.writerows(all_entries)
    
    print(f"\n✅ {len(all_entries)} entregas → {csv_path}")
    print(f"   (cada fila = 1 paquete)")
    
    # Mostrar preview
    print(f"\n── Preview (primeras 10 filas) ──")
    for i, e in enumerate(all_entries[:10]):
        print(f"  {i+1}. {e['cliente'][:25]:<25} | {e['direccion'][:40]:<40} | {e['ciudad']}")


if __name__ == "__main__":
    main()
